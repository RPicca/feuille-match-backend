# backend/app.py basé sur le CSV avec entête: Pseudo;Nom;Licence;Numero
from flask import Flask, request, send_file, jsonify
import openpyxl
import os
from io import BytesIO
import csv
from flask_cors import CORS

app = Flask(__name__)
CORS(app, expose_headers=["Content-Disposition"])
# Constantes Excel
coord_locaux_visiteurs = [[5, 5], [28, 5]]
coord_couleurs = [[6, 9], [29, 9]]
prem_ligne_licences = [15, 38]
col_joueur = 5
col_num = 10

dossier = os.path.dirname(os.path.abspath(__file__))
fichier_licencies = os.path.join(dossier, "licencies.csv")
template_feuille = os.path.join(dossier, "feuille_match_vide.xlsx")

# Chargement des licenciés

def charger_licencies():
    dic_licencies = {}
    with open(fichier_licencies, newline='', encoding='utf-8') as csvfile:
        licencies = csv.reader(csvfile, delimiter=';', quotechar='|')
        next(licencies)  # sauter l'entête
        for l in licencies:
            if len(l) >= 4:
                pseudo = l[0]
                dic_licencies[pseudo] = l[1:]  # [Nom, Licence, Numero]
    return dic_licencies

@app.route('/joueurs', methods=['GET'])
def get_joueurs():
    dic = charger_licencies()
    return jsonify(sorted(dic.keys()))

@app.route('/generate', methods=['POST'])
def generate():
    data = request.json
    equipe = data.get("equipe", "")
    adversaire = data.get("adversaire", "")
    couleur = data.get("couleur", "")
    locaux = data.get("locaux", True)
    pseudos_selectionnes = data.get("joueurs", [])

    dic_licencies = charger_licencies()
    joueurs_selectionnes = {pseudo: dic_licencies[pseudo] for pseudo in pseudos_selectionnes if pseudo in dic_licencies}

    index_locaux = int(not locaux)
    ligne_joueurs = prem_ligne_licences[index_locaux]

    wb = openpyxl.load_workbook(template_feuille)
    feuille = wb.active

    # Écriture des équipes
    feuille.cell(coord_locaux_visiteurs[index_locaux][0], coord_locaux_visiteurs[index_locaux][1], equipe)
    feuille.cell(coord_locaux_visiteurs[index_locaux-1][0], coord_locaux_visiteurs[index_locaux-1][1], adversaire)

    # Écriture couleur
    feuille.cell(coord_couleurs[index_locaux][0], coord_couleurs[index_locaux][1], couleur)

    # Écriture joueurs
    for pseudo in joueurs_selectionnes:
        data = joueurs_selectionnes[pseudo]  # [Nom, Licence, Numero]
        feuille.cell(ligne_joueurs, 1, data[1])       # Licence
        feuille.cell(ligne_joueurs, col_joueur, data[0])  # Nom
        feuille.cell(ligne_joueurs, col_num, data[2])     # Numéro
        ligne_joueurs += 1

    output = BytesIO()
    nom_fichier = f"Feuille_{equipe}_{adversaire}.xlsx".replace(" ", "_")
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=nom_fichier, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
